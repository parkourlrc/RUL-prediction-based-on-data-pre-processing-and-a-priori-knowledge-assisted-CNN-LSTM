#!/usr/bin/env python
# -*- coding: utf-8 -*-


from PyEMD import CEEMDAN
import numpy as np 
np.set_printoptions(threshold=np.inf) #保证很大的数组也可以显示全
import matplotlib.pyplot as plt
import pandas as pd
import xlwt
import xlrd
from openpyxl import Workbook
#xlwt,xlrd是python将数据导入excel表格使用的库



# tips：记得设置全局变量 IImfs=[] 
 
def ceemdan_decompose(data):
    ceemdan = CEEMDAN.CEEMDAN()
    c_imfs=ceemdan.ceemdan(data)
    #imfs, res = ceemdan.get_imfs_and_residue()
    imfs, res = c_imfs[:-1],c_imfs[-1]
    plt.figure(figsize=(12,9))
    plt.subplots_adjust(hspace=0.1)
    plt.subplot(imfs.shape[0]+2, 1, 1)
    plt.plot(data,'r')
    for i in range(imfs.shape[0]):
        plt.subplot(imfs.shape[0]+2,1,i+2)
        plt.plot(imfs[i], 'g')
        plt.ylabel("IMF %i" %(i+1))
        plt.locator_params(axis='x', nbins=10)
        plt.show()
        # 在函数前必须设置一个全局变量 IImfs=[]
        IImfs.append(imfs[i])


# 生成res的分解
def ceemdan_decompose_res(data):
    ceemdan = CEEMDAN.CEEMDAN()
    c_imfs =ceemdan.ceemdan(data)
    #imfs, res = ceemdan.get_imfs_and_residue()
    imfs, res = c_imfs[:-1],c_imfs[-1]
    plt.figure(figsize=(12,9))
    plt.subplots_adjust(hspace=0.1)
    plt.subplot(imfs.shape[0]+3, 1, 1)
    plt.plot(data,'r')
    # 打开工作簿
    wb = Workbook()
    # 选择活动的工作表
    sheet = wb.active
    
    
    for i in range(imfs.shape[0]):
        plt.subplot(imfs.shape[0]+3,1,i+2)
        plt.plot(imfs[i], 'g')
        plt.ylabel("IMF %i" %(i+1))
        plt.locator_params(axis='x', nbins=10)
        # 在函数前必须设置一个全局变量 IImfs=[]
        IImfs.append(imfs[i])
        for j in range(len(imfs[i])):
            sheet.cell(row=j+1, column=i+1).value = imfs[i][j]
    # 保存工作簿
    wb.save("example.xls")
    plt.subplot(imfs.shape[0]+3, 1, imfs.shape[0]+3)
    plt.plot(res,'g')
    plt.show()
    return res


# ceemdan分解
IImfs=[]
records=[]
records=pd.read_excel("B18rongliang.xls")
#records=pd.read_csv("./B18rongliang.csv",'r',encoding='gb18030')
#with open('B18rongliang.csv','r',encoding="gb18030",errors="ignore") as f:
#    import csv
#    reader = csv.reader((line.replace('\0', '') for line in f), delimiter=",")
#    for line in reader:
#        records.append(line)
res=ceemdan_decompose_res(np.array(records).ravel())
#print(res)
#res=ceemdan_decompose_res(D:\竞赛、项目\LSTM预测电池论文\CEEMDAN及其他算法\ceemdan_python\B18rongliang.csv)


out = open("output.txt", "w")

for i in res:
    out.write(str(i) + "\n")

out.close()

'''
# 打开工作簿
wb = Workbook()

# 选择活动的工作表
sheet = wb.active

for i in range(len(array1)):
    sheet.cell(row=i+1, column=1).value = array1[i]
    sheet.cell(row=i+1, column=2).value = array2[i]
    sheet.cell(row=i+1, column=3).value = array3[i]

# 保存工作簿
wb.save("example.xls")
'''

'''
wb = Workbook()

ws = wb.create_sheet("quzao")
#res = [[0],
#         [1],
#         [2],
#         [3]
#         ]
#feature = [
#           [0.1, 0.2, 0.3, 0.4, 0.5],
#           [0.11, 0.21, 0.31, 0.41, 0.51],
#           [0.6, 0.7, 0.8, 0.9, 1.00],
#           [1.1, 1.2, 1.3, 1.4, 1.5],
#           ]
#这个地方之所以 变成numpy格式是因为在很多时候我们都是在numpy格式下计算的，模拟一下预处理
res = np.array(res)
#feature = np.array(feature)

res_input = []
for l in range(len(res)):
    #res_input.append(res[l][0])
    res_input.append(res[l])



ws.append(res_input)
#for f in range(len(feature[0])):

#    ws.append(feature[:, f].tolist())


wb.save("quzaohou.xlsx")
'''
'''
for i in range(len(res)):
    res[i]=res[i][i]

wb = xlwt.Workbook()
# 添加一个表
ws = wb.add_sheet('test')

#添加数据使用.write函数（横坐标，纵坐标，内容）注意横纵坐标从0开始，横纵坐标即对于excel而言,如果数据较多，超过65536就会报错
#ws.write(x坐标, y坐标, 导出内容)
#例：ws.write(0, 0, '股票编号')
#ws.write(0, 0, a),a是变量
ws.write(0, 0, res)

#将数据导出，保存格式为xxx.xls。其中xxx与上文表名同步
#注意此处一定要保存为.xls形式，如果是xlsx会打不开
wb.save('quzhaohou.xls')
'''
